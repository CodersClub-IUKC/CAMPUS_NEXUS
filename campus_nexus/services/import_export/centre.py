from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.exceptions import PermissionDenied, ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone

from campus_nexus.countries import COUNTRY_LOOKUP
from campus_nexus.models import (
    Announcement,
    Association,
    Bill,
    BillableItem,
    BillMembership,
    Charge,
    Course,
    Event,
    Expense,
    Faculty,
    Fee,
    Member,
    Membership,
    Payment,
)
from campus_nexus.services.audit import record_audit_event
from campus_nexus.services.charges import create_charge_custom, get_or_create_charge_for_fee


MAX_PREVIEW_ROWS = 2000
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


class ImportExportError(Exception):
    pass


@dataclass(frozen=True)
class ModuleSpec:
    key: str
    label: str
    section: str
    model: type
    fields: tuple[str, ...]
    required: tuple[str, ...]
    natural_key: tuple[str, ...]
    importable: bool = True
    updateable: bool = True
    financial: bool = False
    system_generated: bool = False


MODULES = {
    "members": ModuleSpec(
        "members",
        "Members",
        "Membership Data",
        Member,
        (
            "registration_number",
            "first_name",
            "last_name",
            "email",
            "phone",
            "member_type",
            "national_id_number",
            "nationality",
            "faculty_name",
            "course_name",
        ),
        ("first_name", "last_name", "email", "phone", "member_type"),
        ("registration_number", "email"),
    ),
    "memberships": ModuleSpec(
        "memberships",
        "Memberships",
        "Membership Data",
        Membership,
        ("member_registration_number", "member_email", "association_name", "status", "subscription_anchor_date"),
        ("association_name",),
        ("member_registration_number", "member_email", "association_name"),
    ),
    "faculties": ModuleSpec(
        "faculties",
        "Faculties",
        "Organization Data",
        Faculty,
        ("name",),
        ("name",),
        ("name",),
    ),
    "courses": ModuleSpec(
        "courses",
        "Courses",
        "Organization Data",
        Course,
        ("name", "faculty_name", "duration_years"),
        ("name", "faculty_name", "duration_years"),
        ("name", "faculty_name"),
    ),
    "fees": ModuleSpec(
        "fees",
        "Fees",
        "Financial Data",
        Fee,
        (
            "association_name",
            "fee_type",
            "amount",
            "duration_months",
            "grace_days",
            "max_missed_cycles",
            "allow_installments",
        ),
        ("association_name", "fee_type", "amount"),
        ("association_name", "fee_type"),
        financial=True,
    ),
    "charges": ModuleSpec(
        "charges",
        "Charges",
        "Financial Data",
        Charge,
        (
            "association_name",
            "member_registration_number",
            "purpose",
            "title",
            "amount_due",
            "amount_paid_total",
            "balance",
            "status",
            "due_date",
            "period_start",
            "period_end",
        ),
        (),
        ("association_name", "member_registration_number", "title"),
        importable=False,
        updateable=False,
        financial=True,
        system_generated=True,
    ),
    "payments": ModuleSpec(
        "payments",
        "Payments",
        "Financial Data",
        Payment,
        (
            "association_name",
            "member_registration_number",
            "member_email",
            "fee_type",
            "purpose",
            "title",
            "amount_due",
            "amount_paid",
            "paid_at",
            "payment_method",
            "reference_code",
            "status",
            "note",
        ),
        ("association_name", "amount_paid"),
        ("reference_code",),
        updateable=False,
        financial=True,
    ),
    "expenses": ModuleSpec(
        "expenses",
        "Expenses",
        "Financial Data",
        Expense,
        (
            "association_name",
            "title",
            "category",
            "amount",
            "spent_at",
            "payment_method",
            "payee",
            "reference_code",
            "status",
            "description",
            "note",
        ),
        ("association_name", "title", "amount"),
        ("association_name", "reference_code"),
        financial=True,
    ),
    "billable_items": ModuleSpec(
        "billable_items",
        "Billable Items",
        "Financial Data",
        BillableItem,
        (
            "association_name",
            "name",
            "description",
            "amount",
            "category",
            "is_recurring",
            "recurrence_type",
            "is_active",
        ),
        ("association_name", "name", "amount"),
        ("association_name", "name"),
        financial=True,
    ),
    "bills": ModuleSpec(
        "bills",
        "Bills",
        "Financial Data",
        Bill,
        (
            "association_name",
            "billable_item_name",
            "title",
            "description",
            "amount",
            "due_date",
            "attachment_type",
            "status",
        ),
        ("association_name", "billable_item_name"),
        ("association_name", "title"),
        financial=True,
    ),
    "bill_memberships": ModuleSpec(
        "bill_memberships",
        "Bill Memberships",
        "Financial Data",
        BillMembership,
        (
            "association_name",
            "bill_title",
            "member_registration_number",
            "member_email",
            "amount_due",
            "amount_waived",
            "status",
        ),
        ("association_name", "bill_title", "amount_due"),
        ("association_name", "bill_title", "member_registration_number", "member_email"),
        financial=True,
    ),
    "events": ModuleSpec(
        "events",
        "Events",
        "Communication Data",
        Event,
        ("association_name", "title", "description", "event_date", "venue", "posted_by_registration_number"),
        ("association_name", "title", "description", "event_date", "venue"),
        ("association_name", "title", "event_date"),
    ),
    "announcements": ModuleSpec(
        "announcements",
        "Announcements",
        "Communication Data",
        Announcement,
        ("title", "message", "audience", "association_name", "faculty_name", "is_published"),
        ("title", "message", "audience"),
        ("title",),
    ),
}


ALIASES = {
    "registration_number": ("registration_number", "reg_no", "registration_no", "student_number"),
    "member_registration_number": ("member_registration_number", "registration_number", "reg_no", "registration_no"),
    "member_email": ("member_email", "email", "email_address"),
    "first_name": ("first_name", "firstname", "given_name"),
    "last_name": ("last_name", "lastname", "surname"),
    "email": ("email", "email_address"),
    "phone": ("phone", "phone_number", "mobile"),
    "national_id_number": ("national_id_number", "nin", "national_id"),
    "faculty_name": ("faculty_name", "faculty", "faculty_code"),
    "course_name": ("course_name", "course", "programme", "course_code"),
    "association_name": ("association_name", "association", "association_code"),
    "billable_item_name": ("billable_item_name", "billable_item", "item_name"),
    "bill_title": ("bill_title", "bill", "title"),
    "posted_by_registration_number": ("posted_by_registration_number", "posted_by", "poster_registration_number"),
}


def get_spec(module_key):
    try:
        return MODULES[module_key]
    except KeyError as exc:
        raise ImportExportError("Unsupported import/export module.") from exc


def user_role(request):
    user = request.user
    if user.is_superuser:
        return "superuser"
    if getattr(user, "guild", None):
        return "guild"
    if getattr(user, "association_admin", None):
        return "association_admin"
    if getattr(user, "dean", None):
        return "dean"
    return "staff"


def can_view_module(request, spec):
    if getattr(request, "is_management_command", False):
        return True
    role = user_role(request)
    if role == "superuser":
        return True
    if role == "guild":
        return spec.key not in {"payments", "expenses"}
    if role == "dean":
        return spec.key not in {"payments", "expenses"}
    if role == "association_admin":
        return spec.key not in {"members", "faculties", "courses"}
    return False


def can_import_module(request, spec):
    if getattr(request, "is_management_command", False):
        return spec.importable
    if not spec.importable or user_role(request) == "dean":
        return False
    if request.user.is_superuser:
        return True
    if user_role(request) == "guild":
        return spec.key in {"members", "memberships", "faculties", "courses", "events", "announcements", "fees", "bills", "billable_items", "bill_memberships"}
    if user_role(request) == "association_admin":
        return spec.key in {"memberships", "fees", "payments", "expenses", "billable_items", "bills", "bill_memberships", "events", "announcements"}
    return False


def modules_for_request(request):
    sections = {}
    for spec in MODULES.values():
        if not can_view_module(request, spec):
            continue
        sections.setdefault(spec.section, []).append(
            {
                "key": spec.key,
                "label": spec.label,
                "importable": can_import_module(request, spec),
                "updateable": spec.updateable,
                "system_generated": spec.system_generated,
            }
        )
    return sections


def scoped_queryset(request, spec):
    qs = spec.model.objects.all()
    role = user_role(request)
    assoc_admin = getattr(request.user, "association_admin", None)
    if role != "association_admin":
        return qs
    assoc = assoc_admin.association
    if spec.key in {"memberships", "fees", "expenses", "billable_items", "bills", "events"}:
        return qs.filter(association=assoc)
    if spec.key in {"charges", "payments"}:
        return qs.filter(membership__association=assoc)
    if spec.key == "bill_memberships":
        return qs.filter(bill__association=assoc)
    if spec.key == "announcements":
        return qs.filter(association=assoc)
    return qs.none()


def normalize_header(value):
    value = (value or "").strip().lower()
    value = re.sub(r"[\s\-]+", "_", value)
    value = re.sub(r"[^a-z0-9_]", "", value)
    return value


def normalize_headers(headers, spec):
    aliases = {field: tuple(normalize_header(x) for x in ALIASES.get(field, (field,)) + (field,)) for field in spec.fields}
    reverse = {}
    for field, options in aliases.items():
        for option in options:
            reverse.setdefault(option, set()).add(field)

    mapping = {}
    unknown = []
    ambiguous = []
    for header in headers:
        normalized = normalize_header(header)
        matches = reverse.get(normalized, set())
        if len(matches) > 1:
            ambiguous.append(header)
        elif len(matches) == 1:
            field = next(iter(matches))
            if field in mapping.values():
                ambiguous.append(header)
            else:
                mapping[header] = field
        else:
            unknown.append(header)

    missing = [field for field in spec.required if field not in mapping.values()]
    return mapping, missing, ambiguous, unknown


def parse_upload(uploaded_file):
    name = uploaded_file.name or ""
    ext = "." + name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if ext not in SUPPORTED_EXTENSIONS:
        raise ImportExportError("Upload a .csv or .xlsx file.")

    if ext == ".csv":
        text = uploaded_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ImportExportError("The file has no header row.")
        return ext, reader.fieldnames, list(reader)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportExportError("Excel import requires openpyxl to be installed.") from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ImportExportError("The workbook is empty.")
    headers = [str(value or "").strip() for value in rows[0]]
    data = []
    for values in rows[1:]:
        data.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    return ext, headers, data


def _value(row, field):
    value = row.get(field, "")
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _str(row, field):
    value = _value(row, field)
    return str(value).strip() if value is not None else ""


def _decimal(row, field, required=False):
    value = _str(row, field)
    if not value:
        if required:
            raise ValidationError({field: "This field is required."})
        return None
    try:
        amount = Decimal(value.replace(",", ""))
    except (InvalidOperation, AttributeError):
        raise ValidationError({field: "Enter a valid decimal amount."})
    return amount


def _int(row, field, default=None):
    value = _str(row, field)
    if not value:
        return default
    try:
        return int(value)
    except ValueError:
        raise ValidationError({field: "Enter a valid integer."})


def _bool(row, field, default=False):
    value = _str(row, field).lower()
    if not value:
        return default
    if value in {"1", "true", "yes", "y", "active", "published"}:
        return True
    if value in {"0", "false", "no", "n", "inactive", "draft"}:
        return False
    raise ValidationError({field: "Enter true or false."})


def _date(row, field, required=False):
    value = _value(row, field)
    if not value:
        if required:
            raise ValidationError({field: "This field is required."})
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    raise ValidationError({field: "Enter a valid date, preferably YYYY-MM-DD."})


def _datetime(row, field, required=False):
    value = _value(row, field)
    if not value:
        if required:
            raise ValidationError({field: "This field is required."})
        return None
    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value
    if isinstance(value, date):
        return timezone.make_aware(datetime.combine(value, datetime.min.time()))
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(str(value).strip(), fmt)
            return timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed
        except ValueError:
            pass
    raise ValidationError({field: "Enter a valid date/time."})


def _choice(model, field, value, default=None):
    value = (value or default or "").strip()
    field_obj = model._meta.get_field(field)
    valid = {choice[0] for choice in field_obj.choices}
    if value not in valid:
        raise ValidationError({field: f"Use one of: {', '.join(sorted(valid))}."})
    return value


def _one(model, **lookups):
    return model.objects.filter(**lookups).first()


def _association(name, request):
    assoc_admin = getattr(request.user, "association_admin", None)
    if assoc_admin:
        if name and assoc_admin.association.name.lower() != str(name).strip().lower():
            raise ValidationError({"association_name": "Association admins can only import data for their own association."})
        return assoc_admin.association
    if not name:
        raise ValidationError({"association_name": "This field is required."})
    obj = _one(Association, name__iexact=str(name).strip())
    if not obj:
        raise ValidationError({"association_name": f"Association '{name}' was not found."})
    return obj


def _faculty(name, required=False):
    if not name:
        if required:
            raise ValidationError({"faculty_name": "This field is required."})
        return None
    obj = _one(Faculty, name__iexact=str(name).strip())
    if not obj:
        raise ValidationError({"faculty_name": f"Faculty '{name}' was not found."})
    return obj


def _course(name, faculty=None, required=False):
    if not name:
        if required:
            raise ValidationError({"course_name": "This field is required."})
        return None
    qs = Course.objects.filter(name__iexact=str(name).strip())
    if faculty:
        qs = qs.filter(faculty=faculty)
    obj = qs.first()
    if not obj:
        raise ValidationError({"course_name": f"Course '{name}' was not found."})
    return obj


def _member(row):
    reg = _str(row, "member_registration_number") or _str(row, "registration_number")
    email = _str(row, "member_email") or _str(row, "email")
    qs = Member.objects.all()
    if reg:
        obj = qs.filter(registration_number__iexact=reg).first()
    elif email:
        obj = qs.filter(email__iexact=email).first()
    else:
        raise ValidationError({"member_registration_number": "Provide member_registration_number or member_email."})
    if not obj:
        raise ValidationError({"member_registration_number": "Member was not found."})
    return obj


def _membership(row, request, required_assoc=True):
    assoc = _association(_str(row, "association_name"), request) if required_assoc else None
    member = _member(row)
    qs = Membership.objects.filter(member=member)
    if assoc:
        qs = qs.filter(association=assoc)
    obj = qs.first()
    if not obj:
        raise ValidationError({"member_registration_number": "Membership was not found for this association."})
    return obj


def _clean_error(exc):
    if hasattr(exc, "message_dict"):
        return "; ".join(f"{field}: {', '.join(messages)}" for field, messages in exc.message_dict.items())
    if hasattr(exc, "messages"):
        return "; ".join(str(message) for message in exc.messages)
    return str(exc)


def _row_data(raw_row, mapping):
    normalized = {}
    for header, value in raw_row.items():
        field = mapping.get(header)
        if field:
            normalized[field] = value
    return normalized


def _existing_for(spec, row, request):
    if spec.key == "members":
        reg = _str(row, "registration_number")
        email = _str(row, "email")
        return (Member.objects.filter(registration_number__iexact=reg).first() if reg else None) or (
            Member.objects.filter(email__iexact=email).first() if email else None
        )
    if spec.key == "memberships":
        try:
            member = _member(row)
            association = _association(_str(row, "association_name"), request)
        except ValidationError:
            return None
        return Membership.objects.filter(member=member, association=association).first()
    if spec.key == "faculties":
        return Faculty.objects.filter(name__iexact=_str(row, "name")).first()
    if spec.key == "courses":
        faculty = Faculty.objects.filter(name__iexact=_str(row, "faculty_name")).first()
        return Course.objects.filter(name__iexact=_str(row, "name"), faculty=faculty).first() if faculty else None
    if spec.key == "fees":
        association = Association.objects.filter(name__iexact=_str(row, "association_name")).first()
        return Fee.objects.filter(association=association, fee_type=_str(row, "fee_type")).first() if association else None
    if spec.key == "payments":
        ref = _str(row, "reference_code")
        if ref:
            return Payment.objects.filter(reference_code__iexact=ref).first()
        return None
    if spec.key == "expenses":
        ref = _str(row, "reference_code")
        association = Association.objects.filter(name__iexact=_str(row, "association_name")).first()
        if ref and association:
            return Expense.objects.filter(association=association, reference_code__iexact=ref).first()
        return None
    if spec.key == "billable_items":
        association = Association.objects.filter(name__iexact=_str(row, "association_name")).first()
        return BillableItem.objects.filter(association=association, name__iexact=_str(row, "name")).first() if association else None
    if spec.key == "bills":
        association = Association.objects.filter(name__iexact=_str(row, "association_name")).first()
        title = _str(row, "title") or _str(row, "billable_item_name")
        return Bill.objects.filter(association=association, title__iexact=title).first() if association and title else None
    if spec.key == "bill_memberships":
        try:
            association = _association(_str(row, "association_name"), request)
            bill = Bill.objects.filter(association=association, title__iexact=_str(row, "bill_title")).first()
            member = _member(row)
            membership = Membership.objects.filter(association=association, member=member).first()
        except ValidationError:
            return None
        return BillMembership.objects.filter(bill=bill, membership=membership).first() if bill and membership else None
    if spec.key == "events":
        association = Association.objects.filter(name__iexact=_str(row, "association_name")).first()
        return Event.objects.filter(association=association, title__iexact=_str(row, "title"), event_date=_datetime(row, "event_date")).first() if association and _str(row, "event_date") else None
    if spec.key == "announcements":
        return Announcement.objects.filter(title__iexact=_str(row, "title")).first()
    return None


def _build_object(spec, row, request, existing=None):
    if spec.key == "members":
        faculty = _faculty(_str(row, "faculty_name"))
        course = _course(_str(row, "course_name"), faculty=faculty)
        email = _str(row, "email").lower()
        if email:
            validate_email(email)
        obj = existing or Member()
        obj.first_name = _str(row, "first_name")
        obj.last_name = _str(row, "last_name")
        obj.email = email
        obj.phone = _str(row, "phone").replace(" ", "").replace("-", "")
        if obj.phone.startswith("00"):
            obj.phone = "+" + obj.phone[2:]
        obj.registration_number = _str(row, "registration_number") or None
        obj.national_id_number = _str(row, "national_id_number") or None
        obj.member_type = _choice(Member, "member_type", _str(row, "member_type"))
        nationality = _str(row, "nationality")
        obj.nationality = COUNTRY_LOOKUP.get(nationality.lower(), nationality)
        obj.faculty = faculty
        obj.course = course
        if not existing and getattr(request.user, "pk", None):
            obj.created_by = request.user
            assoc_admin = getattr(request.user, "association_admin", None)
            if assoc_admin:
                obj.created_in_association = assoc_admin.association
        return obj

    if spec.key == "memberships":
        member = _member(row)
        association = _association(_str(row, "association_name"), request)
        obj = existing or Membership(member=member, association=association)
        obj.status = _choice(Membership, "status", _str(row, "status"), default="active")
        anchor = _date(row, "subscription_anchor_date")
        if anchor:
            obj.subscription_anchor_date = anchor
        return obj

    if spec.key == "faculties":
        obj = existing or Faculty()
        obj.name = _str(row, "name")
        return obj

    if spec.key == "courses":
        obj = existing or Course()
        obj.name = _str(row, "name")
        obj.faculty = _faculty(_str(row, "faculty_name"), required=True)
        obj.duration_years = _int(row, "duration_years")
        return obj

    if spec.key == "fees":
        obj = existing or Fee()
        obj.association = _association(_str(row, "association_name"), request)
        obj.fee_type = _choice(Fee, "fee_type", _str(row, "fee_type"))
        obj.amount = _decimal(row, "amount", required=True)
        obj.duration_months = _int(row, "duration_months", 0) or 0
        obj.grace_days = _int(row, "grace_days", 0) or 0
        obj.max_missed_cycles = _int(row, "max_missed_cycles", 2) or 0
        obj.allow_installments = _bool(row, "allow_installments", True)
        return obj

    if spec.key == "payments":
        return make_payment_from_row(row, request)

    if spec.key == "expenses":
        obj = existing or Expense()
        obj.association = _association(_str(row, "association_name"), request)
        obj.title = _str(row, "title")
        obj.category = _choice(Expense, "category", _str(row, "category"), default="other")
        obj.amount = _decimal(row, "amount", required=True)
        obj.spent_at = _datetime(row, "spent_at") or timezone.now()
        obj.payment_method = _choice(Expense, "payment_method", _str(row, "payment_method"), default="cash")
        obj.payee = _str(row, "payee")
        obj.reference_code = _str(row, "reference_code")
        obj.status = _choice(Expense, "status", _str(row, "status"), default="recorded")
        obj.description = _str(row, "description")
        obj.note = _str(row, "note")
        if not existing:
            obj.recorded_by = request.user
        return obj

    if spec.key == "billable_items":
        obj = existing or BillableItem()
        obj.association = _association(_str(row, "association_name"), request)
        obj.name = _str(row, "name")
        obj.description = _str(row, "description")
        obj.amount = _decimal(row, "amount", required=True)
        obj.category = _choice(BillableItem, "category", _str(row, "category"), default="other")
        obj.is_recurring = _bool(row, "is_recurring", False)
        recurrence = _str(row, "recurrence_type")
        obj.recurrence_type = recurrence or ("one_time" if obj.is_recurring else None)
        obj.is_active = _bool(row, "is_active", True)
        if not existing:
            obj.created_by = request.user
        return obj

    if spec.key == "bills":
        association = _association(_str(row, "association_name"), request)
        item = BillableItem.objects.filter(association=association, name__iexact=_str(row, "billable_item_name")).first()
        if not item:
            raise ValidationError({"billable_item_name": "Billable item was not found for this association."})
        obj = existing or Bill()
        obj.association = association
        obj.billable_item = item
        obj.title = _str(row, "title") or item.name
        obj.description = _str(row, "description")
        obj.amount = _decimal(row, "amount") or item.amount
        obj.due_date = _date(row, "due_date")
        obj.attachment_type = _choice(Bill, "attachment_type", _str(row, "attachment_type"), default="selective")
        obj.status = _choice(Bill, "status", _str(row, "status"), default="draft")
        if not existing:
            obj.created_by = request.user
        return obj

    if spec.key == "bill_memberships":
        association = _association(_str(row, "association_name"), request)
        bill = Bill.objects.filter(association=association, title__iexact=_str(row, "bill_title")).first()
        if not bill:
            raise ValidationError({"bill_title": "Bill was not found for this association."})
        membership = _membership(row, request)
        if membership.association_id != association.id:
            raise ValidationError({"member_registration_number": "Membership belongs to a different association."})
        obj = existing or BillMembership(bill=bill, membership=membership)
        obj.amount_due = _decimal(row, "amount_due", required=True)
        obj.amount_waived = _decimal(row, "amount_waived") or Decimal("0")
        obj.status = _choice(BillMembership, "status", _str(row, "status"), default="unpaid")
        return obj

    if spec.key == "events":
        association = _association(_str(row, "association_name"), request)
        obj = existing or Event()
        obj.association = association
        obj.title = _str(row, "title")
        obj.description = _str(row, "description")
        obj.event_date = _datetime(row, "event_date", required=True)
        obj.venue = _str(row, "venue")
        posted_reg = _str(row, "posted_by_registration_number")
        if posted_reg:
            member = Member.objects.filter(registration_number__iexact=posted_reg).first()
            obj.posted_by = Membership.objects.filter(member=member, association=association).first() if member else None
            if not obj.posted_by:
                raise ValidationError({"posted_by_registration_number": "Posting membership was not found."})
        return obj

    if spec.key == "announcements":
        obj = existing or Announcement()
        obj.title = _str(row, "title")
        obj.message = _str(row, "message")
        obj.audience = _choice(Announcement, "audience", _str(row, "audience"), default="all")
        obj.association = _association(_str(row, "association_name"), request) if obj.audience == "association" else None
        obj.faculty = _faculty(_str(row, "faculty_name"), required=obj.audience == "faculty") if obj.audience == "faculty" else None
        obj.is_published = _bool(row, "is_published", True)
        if not existing:
            obj.posted_by = request.user
        return obj

    raise ImportExportError("This module does not support imports.")


def build_preview(request, module_key, uploaded_file, mode):
    spec = get_spec(module_key)
    if not can_import_module(request, spec):
        raise PermissionDenied("You are not allowed to import this module.")
    if mode not in {"create", "upsert"}:
        mode = "create"
    if mode == "upsert" and not spec.updateable:
        mode = "create"

    ext, headers, raw_rows = parse_upload(uploaded_file)
    if len(raw_rows) > MAX_PREVIEW_ROWS:
        raise ImportExportError(f"Import files are limited to {MAX_PREVIEW_ROWS} rows per upload.")
    mapping, missing, ambiguous, unknown = normalize_headers(headers, spec)
    if ambiguous:
        raise ImportExportError("Ambiguous column(s): " + ", ".join(ambiguous))
    if missing:
        raise ImportExportError("Missing required column(s): " + ", ".join(missing))

    seen = set()
    rows = []
    for index, raw_row in enumerate(raw_rows, start=2):
        normalized = _row_data(raw_row, mapping)
        status = "CREATE"
        error = ""
        existing = None
        try:
            natural = tuple(str(normalized.get(field, "")).strip().lower() for field in spec.natural_key if normalized.get(field, ""))
            if natural and natural in seen:
                status = "DUPLICATE"
                error = "Duplicate row in uploaded file."
            else:
                if natural:
                    seen.add(natural)
                existing = _existing_for(spec, normalized, request)
                if existing:
                    status = "UPDATE" if mode == "upsert" and spec.updateable else "DUPLICATE"
                    if status == "DUPLICATE":
                        error = "Matching record already exists."
                obj = _build_object(spec, normalized, request, existing if status == "UPDATE" else None)
                obj.full_clean()
        except Exception as exc:
            if status != "DUPLICATE":
                status = "INVALID"
            error = error or _clean_error(exc)

        rows.append(
            {
                "row_number": index,
                "data": normalized,
                "status": status,
                "error": error,
            }
        )

    counts = {
        "total": len(rows),
        "valid": sum(1 for row in rows if row["status"] in {"CREATE", "UPDATE"}),
        "invalid": sum(1 for row in rows if row["status"] == "INVALID"),
        "duplicates": sum(1 for row in rows if row["status"] == "DUPLICATE"),
        "creates": sum(1 for row in rows if row["status"] == "CREATE"),
        "updates": sum(1 for row in rows if row["status"] == "UPDATE"),
    }
    return {
        "module": spec.key,
        "module_label": spec.label,
        "file_name": uploaded_file.name,
        "file_type": ext,
        "mode": mode,
        "headers": spec.fields,
        "unknown_headers": unknown,
        "rows": rows,
        "counts": counts,
    }


@transaction.atomic
def commit_preview(request, preview):
    spec = get_spec(preview["module"])
    if not can_import_module(request, spec):
        raise PermissionDenied("You are not allowed to import this module.")

    created = updated = skipped = 0
    for row in preview["rows"]:
        if row["status"] not in {"CREATE", "UPDATE"}:
            skipped += 1
            continue
        existing = _existing_for(spec, row["data"], request)
        if row["status"] == "UPDATE":
            obj = _build_object(spec, row["data"], request, existing)
        else:
            obj = _build_object(spec, row["data"], request, None)
        obj.full_clean()
        obj.save()

        if spec.key == "payments":
            _finalize_payment_import(obj, row["data"], request)
        elif spec.key == "bill_memberships":
            obj.update_status_from_payments()

        record_audit_event(
            actor=request.user,
            action=f"{spec.key[:-1] if spec.key.endswith('s') else spec.key}_imported",
            obj=obj,
            metadata={"import_status": row["status"], "module": spec.key},
        )
        if row["status"] == "UPDATE":
            updated += 1
        else:
            created += 1

    return {"created": created, "updated": updated, "skipped": skipped}


def _finalize_payment_import(payment, row, request):
    if payment.recorded_by_id is None:
        payment.recorded_by = request.user
    membership = payment.membership
    fee_type = _str(row, "fee_type")
    fee = Fee.objects.filter(association=membership.association, fee_type=fee_type).first() if fee_type else None
    payment.fee = fee
    if fee:
        charge = get_or_create_charge_for_fee(membership=membership, fee=fee, user=request.user)
    else:
        amount_due = _decimal(row, "amount_due") or payment.amount_paid
        charge = create_charge_custom(
            membership=membership,
            purpose=_choice(Charge, "purpose", _str(row, "purpose"), default="other"),
            title=_str(row, "title") or "Imported payment",
            amount_due=amount_due,
            due_date=None,
            description=_str(row, "note"),
            user=request.user,
        )
    payment.charge = charge
    payment.save()
    charge.recompute_status()
    charge.save(update_fields=["status"])
    bill_membership = getattr(charge, "bill_membership", None)
    if bill_membership:
        bill_membership.update_status_from_payments()


def make_payment_from_row(row, request):
    membership = _membership(row, request)
    fee_type = _str(row, "fee_type")
    if fee_type and not Fee.objects.filter(association=membership.association, fee_type=fee_type).exists():
        raise ValidationError({"fee_type": "Fee was not found for this association."})
    obj = Payment()
    obj.membership = membership
    obj.amount_paid = _decimal(row, "amount_paid", required=True)
    obj.paid_at = _datetime(row, "paid_at") or timezone.now()
    obj.payment_method = _choice(Payment, "payment_method", _str(row, "payment_method"), default="cash")
    obj.reference_code = _str(row, "reference_code")
    obj.status = _choice(Payment, "status", _str(row, "status"), default="recorded")
    obj.note = _str(row, "note")
    obj.recorded_by = request.user
    return obj


def preview_to_session(preview):
    return json.dumps(preview, default=str)


def preview_from_session(value):
    return json.loads(value)


def export_rows(request, spec):
    qs = scoped_queryset(request, spec)
    if spec.key == "members":
        qs = qs.select_related("faculty", "course")
        return [
            {
                "registration_number": obj.registration_number,
                "first_name": obj.first_name,
                "last_name": obj.last_name,
                "email": obj.email,
                "phone": obj.phone,
                "member_type": obj.member_type,
                "national_id_number": obj.national_id_number,
                "nationality": obj.nationality,
                "faculty_name": obj.faculty.name if obj.faculty else "",
                "course_name": obj.course.name if obj.course else "",
            }
            for obj in qs
        ]
    if spec.key == "memberships":
        qs = qs.select_related("member", "association")
        return [
            {
                "member_registration_number": obj.member.registration_number,
                "member_email": obj.member.email,
                "association_name": obj.association.name,
                "status": obj.status,
                "subscription_anchor_date": obj.subscription_anchor_date,
            }
            for obj in qs
        ]
    if spec.key == "faculties":
        return [{"name": obj.name} for obj in qs]
    if spec.key == "courses":
        return [{"name": obj.name, "faculty_name": obj.faculty.name, "duration_years": obj.duration_years} for obj in qs.select_related("faculty")]
    if spec.key == "fees":
        return [
            {
                "association_name": obj.association.name,
                "fee_type": obj.fee_type,
                "amount": obj.amount,
                "duration_months": obj.duration_months,
                "grace_days": obj.grace_days,
                "max_missed_cycles": obj.max_missed_cycles,
                "allow_installments": obj.allow_installments,
            }
            for obj in qs.select_related("association")
        ]
    if spec.key == "charges":
        return [
            {
                "association_name": obj.association.name,
                "member_registration_number": obj.membership.member.registration_number,
                "purpose": obj.purpose,
                "title": obj.title,
                "amount_due": obj.amount_due,
                "amount_paid_total": obj.amount_paid_total,
                "balance": obj.balance,
                "status": obj.status,
                "due_date": obj.due_date,
                "period_start": obj.period_start,
                "period_end": obj.period_end,
            }
            for obj in qs.select_related("association", "membership__member")
        ]
    if spec.key == "payments":
        return [
            {
                "association_name": obj.membership.association.name,
                "member_registration_number": obj.membership.member.registration_number,
                "member_email": obj.membership.member.email,
                "fee_type": obj.fee.fee_type if obj.fee else "",
                "purpose": obj.charge.purpose if obj.charge else "",
                "title": obj.charge.title if obj.charge else "",
                "amount_due": obj.charge.amount_due if obj.charge else "",
                "amount_paid": obj.amount_paid,
                "paid_at": obj.paid_at,
                "payment_method": obj.payment_method,
                "reference_code": obj.reference_code,
                "status": obj.status,
                "note": obj.note,
            }
            for obj in qs.select_related("membership__association", "membership__member", "fee", "charge")
        ]
    if spec.key == "expenses":
        return [
            {
                "association_name": obj.association.name,
                "title": obj.title,
                "category": obj.category,
                "amount": obj.amount,
                "spent_at": obj.spent_at,
                "payment_method": obj.payment_method,
                "payee": obj.payee,
                "reference_code": obj.reference_code,
                "status": obj.status,
                "description": obj.description,
                "note": obj.note,
            }
            for obj in qs.select_related("association")
        ]
    if spec.key == "billable_items":
        return [
            {
                "association_name": obj.association.name,
                "name": obj.name,
                "description": obj.description,
                "amount": obj.amount,
                "category": obj.category,
                "is_recurring": obj.is_recurring,
                "recurrence_type": obj.recurrence_type,
                "is_active": obj.is_active,
            }
            for obj in qs.select_related("association")
        ]
    if spec.key == "bills":
        return [
            {
                "association_name": obj.association.name,
                "billable_item_name": obj.billable_item.name,
                "title": obj.title,
                "description": obj.description,
                "amount": obj.amount,
                "due_date": obj.due_date,
                "attachment_type": obj.attachment_type,
                "status": obj.status,
            }
            for obj in qs.select_related("association", "billable_item")
        ]
    if spec.key == "bill_memberships":
        return [
            {
                "association_name": obj.bill.association.name,
                "bill_title": obj.bill.title,
                "member_registration_number": obj.membership.member.registration_number,
                "member_email": obj.membership.member.email,
                "amount_due": obj.amount_due,
                "amount_waived": obj.amount_waived,
                "status": obj.status,
            }
            for obj in qs.select_related("bill__association", "membership__member")
        ]
    if spec.key == "events":
        return [
            {
                "association_name": obj.association.name,
                "title": obj.title,
                "description": obj.description,
                "event_date": obj.event_date,
                "venue": obj.venue,
                "posted_by_registration_number": obj.posted_by.member.registration_number if obj.posted_by else "",
            }
            for obj in qs.select_related("association", "posted_by__member")
        ]
    if spec.key == "announcements":
        return [
            {
                "title": obj.title,
                "message": obj.message,
                "audience": obj.audience,
                "association_name": obj.association.name if obj.association else "",
                "faculty_name": obj.faculty.name if obj.faculty else "",
                "is_published": obj.is_published,
            }
            for obj in qs.select_related("association", "faculty")
        ]
    return []


def export_csv_response(request, module_key, template=False):
    spec = get_spec(module_key)
    if not can_view_module(request, spec):
        raise PermissionDenied("You are not allowed to export this module.")
    rows = [] if template else export_rows(request, spec)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    suffix = "template" if template else timezone.localdate().isoformat()
    response["Content-Disposition"] = f'attachment; filename="campus_nexus_{spec.key}_{suffix}.csv"'
    writer = csv.DictWriter(response, fieldnames=spec.fields)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return response


def export_excel_response(request, module_key, template=False):
    spec = get_spec(module_key)
    if not can_view_module(request, spec):
        raise PermissionDenied("You are not allowed to export this module.")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportExportError("Excel export requires openpyxl to be installed.") from exc

    rows = [] if template else export_rows(request, spec)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec.label[:31]
    sheet.append(list(spec.fields))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for row in rows:
        sheet.append([row.get(field, "") for field in spec.fields])
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 42)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    suffix = "template" if template else timezone.localdate().isoformat()
    response["Content-Disposition"] = f'attachment; filename="campus_nexus_{spec.key}_{suffix}.xlsx"'
    workbook.save(response)
    return response


def build_payment_for_validation(row, request):
    return make_payment_from_row(row, request)
